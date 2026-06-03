"""
screen_reports.py — Reports Hub (Phase 3)
8 reports with date range filter, treeview display, Excel export.
Designed for 60+ age users: large buttons, clear layout.
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import date, timedelta
import os

from config import COLORS, FONTS
from ui_utils import place_popup

# ── Report catalogue ──────────────────────────────────────────
REPORTS = [
    {"key": "daily_sales",     "title": "Daily Sales",      "emoji": "📅",
     "color": COLORS["btn_primary"], "needs_dates": True,  "needs_cust": False,
     "cols": [("date","Date",120),("bills","Bills",70),("subtotal","Subtotal ₹",130),
               ("discount","Discount ₹",120),("total","Total ₹",130)],
     "summary_col": "total"},
    {"key": "itemwise",        "title": "Item-wise Sales",  "emoji": "📦",
     "color": COLORS["btn_success"], "needs_dates": True,  "needs_cust": False,
     "cols": [("product_name","Product",220),("qty_sold","Qty",80),
               ("avg_price","Avg Price ₹",120),("discount","Discount ₹",110),
               ("total_sales","Total Sales ₹",140)],
     "summary_col": "total_sales"},
    {"key": "top_products",    "title": "Top Products",     "emoji": "🏆",
     "color": COLORS["btn_warning"], "needs_dates": True,  "needs_cust": False,
     "cols": [("rank","#",50),("product_name","Product",240),
               ("qty_sold","Qty Sold",100),("revenue","Revenue ₹",140)],
     "summary_col": "revenue"},
    {"key": "low_stock",       "title": "Low Stock Alert",  "emoji": "⚠️",
     "color": COLORS["btn_danger"], "needs_dates": False, "needs_cust": False,
     "cols": [("product_code","Code",100),("name","Product",200),
               ("category","Category",130),("stock","Stock",90),
               ("reorder","Reorder",90),("shortage","Shortage",90),
               ("status","Status",120)],
     "summary_col": None},
    {"key": "purchase",        "title": "Purchase / GRN",   "emoji": "🛒",
     "color": COLORS["btn_purple"], "needs_dates": True,  "needs_cust": False,
     "cols": [("date","Date",110),("grn_number","GRN No",120),
               ("supplier_name","Supplier",200),("items","Items",70),
               ("total","Total ₹",130)],
     "summary_col": "total"},
    {"key": "profit_margin",   "title": "Profit & Margin",  "emoji": "💰",
     "color": "#00695C", "needs_dates": True,  "needs_cust": False,
     "cols": [("product_name","Product",200),("qty_sold","Qty",70),
               ("sell_price","Sell ₹",100),("cost_price","Cost ₹",100),
               ("margin_per_unit","Margin ₹",110),("margin_pct","Margin %",100),
               ("total_profit","Total Profit ₹",140)],
     "summary_col": "total_profit"},
    {"key": "stock_valuation", "title": "Stock Valuation",  "emoji": "📋",
     "color": "#1A237E", "needs_dates": False, "needs_cust": False,
     "cols": [("product_code","Code",90),("name","Product",180),
               ("category","Category",120),("stock","Stock",80),
               ("cost_price","Cost ₹",100),("sell_price","Sell ₹",100),
               ("cost_value","Cost Value ₹",130),("retail_value","Retail Value ₹",140)],
     "summary_col": "cost_value"},
    {"key": "customer_ledger", "title": "Customer Ledger",  "emoji": "👥",
     "color": "#4E342E", "needs_dates": True,  "needs_cust": True,
     "cols": [("created_at","Date & Time",160),("customer","Customer",160),
               ("txn_type","Type",110),("amount","Amount ₹",120),
               ("reference","Reference",150),("notes","Notes",160)],
     "summary_col": "amount"},
    {"key": "slow_moving",     "title": "Slow-Moving Items", "emoji": "🐌",
     "color": "#6D4C41", "needs_dates": False, "needs_cust": False,
     "cols": [("product_code","Code",90),("name","Product",200),
               ("category","Category",120),("unit","Unit",70),
               ("current_stock","Stock",90),("selling_price","Price ₹",100),
               ("last_sold","Last Sold",130),("total_qty_30d","Qty(30d)",90)],
     "summary_col": None},
    {"key": "supplier_payables","title": "Supplier Payables","emoji": "🏭",
     "color": "#28A745", "needs_dates": False, "needs_cust": False,
     "cols": [("supplier_name","Supplier",180),("grn_number","GRN",110),
               ("purchase_date","Date",110),("invoice_amount","Invoice ₹",120),
               ("paid_amount","Paid ₹",110),("balance","Balance ₹",120),
               ("age_days","Age (days)",90),("ageing","Bucket",100)],
     "summary_col": "balance"},
    {"key": "customer_ageing",  "title": "Customer Ageing",  "emoji": "📆",
     "color": "#880E4F", "needs_dates": False, "needs_cust": False,
     "cols": [("customer","Customer",180),("phone","Phone",130),
               ("total_due","Total Due ₹",130),("due_0_30","0-30 days ₹",120),
               ("due_31_60","31-60 days ₹",120),("due_60plus","60+ days ₹",120),
               ("last_txn_date","Last Txn",110)],
     "summary_col": "total_due"},
]


class ReportScreen(ctk.CTkFrame):
    def __init__(self, parent, db, current_user, app):
        super().__init__(parent, fg_color=COLORS["bg_main"], corner_radius=0)
        self.db           = db
        self.current_user = current_user
        self.app          = app
        self._active_key  = None
        self._report_data = []
        self._report_def  = None
        self._cust_map    = {}   # name → id for customer ledger
        self._build()

    def _build(self):
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # ── Header ───────────────────────────────────────────
        hdr = ctk.CTkFrame(self, fg_color=COLORS["bg_card"], corner_radius=0, height=70)
        hdr.grid(row=0, column=0, columnspan=2, sticky="ew")
        hdr.grid_propagate(False)
        ctk.CTkLabel(hdr, text="📊   Reports & Analytics",
                     font=FONTS["heading"], text_color=COLORS["text_dark"]
                    ).pack(side="left", padx=25, pady=15)

        # ── Left sidebar: report buttons ─────────────────────
        left = ctk.CTkScrollableFrame(self, fg_color=COLORS["bg_card"],
                                       corner_radius=0, width=220)
        left.grid(row=1, column=0, sticky="nsew")
        ctk.CTkLabel(left, text="SELECT REPORT",
                     font=FONTS["small_bold"], text_color=COLORS["text_muted"]
                    ).pack(pady=(16, 8), padx=14, anchor="w")

        self._rpt_btns = {}
        for rpt in REPORTS:
            btn = ctk.CTkButton(
                left,
                text=f"  {rpt['emoji']}  {rpt['title']}",
                font=FONTS["sidebar"],
                fg_color="transparent",
                hover_color=COLORS["bg_main"],
                text_color=COLORS["text_dark"],
                anchor="w",
                height=52,
                corner_radius=16,
                command=lambda r=rpt: self._select_report(r),
            )
            btn.pack(fill="x", padx=8, pady=3)
            self._rpt_btns[rpt["key"]] = btn

        # ── Right content panel ───────────────────────────────
        right = ctk.CTkFrame(self, fg_color="transparent")
        right.grid(row=1, column=1, sticky="nsew", padx=10, pady=10)
        right.grid_rowconfigure(3, weight=1)
        right.grid_columnconfigure(0, weight=1)

        # Controls bar
        ctrl = ctk.CTkFrame(right, fg_color=COLORS["bg_card"], corner_radius=16, height=68)
        ctrl.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        ctrl.grid_propagate(False)
        ctrl.grid_columnconfigure(99, weight=1)  # push export right

        # Date From
        ctk.CTkLabel(ctrl, text="From:", font=FONTS["small_bold"],
                     text_color=COLORS["text_dark"]).grid(row=0, column=0, padx=(16, 4), pady=14)
        self.date_from = ctk.CTkEntry(ctrl, width=120, font=FONTS["input"],
                                       fg_color=COLORS["bg_input"], border_color=COLORS["border_focus"],
                                       height=40)
        self.date_from.grid(row=0, column=1, padx=(0, 10))
        self.date_from.insert(0, (date.today() - timedelta(days=30)).isoformat())

        # Date To
        ctk.CTkLabel(ctrl, text="To:", font=FONTS["small_bold"],
                     text_color=COLORS["text_dark"]).grid(row=0, column=2, padx=(0, 4))
        self.date_to = ctk.CTkEntry(ctrl, width=120, font=FONTS["input"],
                                     fg_color=COLORS["bg_input"], border_color=COLORS["border_focus"],
                                     height=40)
        self.date_to.grid(row=0, column=3, padx=(0, 10))
        self.date_to.insert(0, date.today().isoformat())

        # Customer selector (hidden unless needed)
        self._cust_frame = ctk.CTkFrame(ctrl, fg_color="transparent")
        self._cust_frame.grid(row=0, column=4, padx=(0, 10))
        ctk.CTkLabel(self._cust_frame, text="Customer:", font=FONTS["small_bold"],
                     text_color=COLORS["text_dark"]).pack(side="left", padx=(0, 4))
        self._cust_var = tk.StringVar(value="All Customers")
        self._cust_menu = ctk.CTkComboBox(self._cust_frame, variable=self._cust_var,
                                           values=["All Customers"], width=170,
                                           font=FONTS["input"], height=40)
        self._cust_menu.pack(side="left")
        self._cust_frame.grid_remove()  # hidden by default

        # Generate button
        ctk.CTkButton(ctrl, text="▶  Generate",
                      font=FONTS["button"], fg_color=COLORS["btn_primary"],
                      height=44, corner_radius=10, width=130,
                      command=self._generate).grid(row=0, column=5, padx=(0, 10))

        # Export buttons
        ctk.CTkButton(ctrl, text="📊 Excel",
                      font=FONTS["small_bold"], fg_color=COLORS["btn_success"],
                      height=40, corner_radius=10, width=90,
                      command=self._export_excel).grid(row=0, column=97, padx=(0, 6))
        ctk.CTkButton(ctrl, text="📋 CSV",
                      font=FONTS["small_bold"], fg_color=COLORS["btn_secondary"],
                      height=40, corner_radius=10, width=80,
                      command=self._export_csv).grid(row=0, column=98, padx=(0, 6))
        ctk.CTkButton(ctrl, text="📄 PDF",
                      font=FONTS["small_bold"], fg_color=COLORS["btn_purple"],
                      height=40, corner_radius=10, width=80,
                      command=self._export_pdf).grid(row=0, column=99, padx=(0, 14))

        # Placeholder / title label
        self._title_lbl = ctk.CTkLabel(right,
                                        text="← Select a report from the left panel",
                                        font=FONTS["subheading"],
                                        text_color=COLORS["text_muted"])
        self._title_lbl.grid(row=1, column=0, pady=(8, 4), sticky="w")

        # Summary row count / grand total label (RPT-1 fix: was missing entirely)
        self._summary_lbl = ctk.CTkLabel(
            right, text="",
            font=FONTS["small_bold"], text_color=COLORS["text_muted"]
        )
        self._summary_lbl.grid(row=2, column=0, sticky="w", pady=(0, 4))

        # Table area
        tbl_frame = ctk.CTkFrame(right, fg_color=COLORS["bg_card"], corner_radius=16)
        tbl_frame.grid(row=3, column=0, sticky="nsew")
        tbl_frame.grid_rowconfigure(0, weight=1)
        tbl_frame.grid_columnconfigure(0, weight=1)

        # Treeview (columns rebuilt per report in _select_report)
        self.tree = ttk.Treeview(
            tbl_frame, columns=[], show="headings",
            style="Rpt.Treeview", selectmode="browse"
        )
        vsb = ttk.Scrollbar(tbl_frame, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(tbl_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew", padx=(4,0), pady=4)
        vsb.grid(row=0, column=1, sticky="ns",  pady=4)
        hsb.grid(row=1, column=0, sticky="ew",  padx=(4,0))
        self.tree.bind("<Double-1>", self._on_tree_drilldown)

    def _select_report(self, rpt: dict):
        """Called when user clicks a report button on the left panel."""
        self._report_def  = rpt
        self._active_key  = rpt["key"]
        self._report_data = []

        # Highlight active button
        for key, btn in self._rpt_btns.items():
            btn.configure(
                fg_color=COLORS["btn_primary"] if key == rpt["key"] else "transparent",
                text_color="white" if key == rpt["key"] else COLORS["text_dark"],
            )

        self._title_lbl.configure(
            text=f"{rpt['emoji']}  {rpt['title']}",
            text_color=rpt["color"]
        )

        # Refresh customer list if needed
        if rpt["needs_cust"]:
            custs = self.db.get_customers(active_only=True)
            self._cust_map = {c["name"]: c["customer_id"] for c in custs}
            names = ["All Customers"] + list(self._cust_map.keys())
            self._cust_menu.configure(values=names)
            self._cust_var.set("All Customers")

        # Show/hide date inputs
        if rpt["needs_dates"]:
            self.date_from.configure(state="normal")
            self.date_to.configure(state="normal")
        else:
            self.date_from.configure(state="disabled")
            self.date_to.configure(state="disabled")

        # Show/hide customer selector
        if rpt["needs_cust"]:
            self._cust_frame.grid()
        else:
            self._cust_frame.grid_remove()

        # Auto-generate immediately
        self._generate()

    def _generate(self):
        if not self._report_def:
            return
        rpt = self._report_def
        df  = self.date_from.get().strip()
        dt  = self.date_to.get().strip()

        try:
            if rpt["needs_dates"]:
                from datetime import date as dt_
                date.fromisoformat(df)
                date.fromisoformat(dt)
        except ValueError:
            messagebox.showwarning("Date Error",
                                   "Enter valid dates in YYYY-MM-DD format.",
                                   parent=self.winfo_toplevel())
            return

        # Fetch data
        try:
            if   rpt["key"] == "daily_sales":     data = self.db.report_daily_sales(df, dt)
            elif rpt["key"] == "itemwise":         data = self.db.report_itemwise_sales(df, dt)
            elif rpt["key"] == "top_products":     data = self.db.report_top_products(df, dt)
            elif rpt["key"] == "low_stock":        data = self.db.report_low_stock()
            elif rpt["key"] == "purchase":         data = self.db.report_purchase(df, dt)
            elif rpt["key"] == "profit_margin":    data = self.db.report_profit_margin(df, dt)
            elif rpt["key"] == "stock_valuation":  data = self.db.report_stock_valuation()
            elif rpt["key"] == "customer_ledger":
                cust_name = self._cust_var.get()
                cust_id   = self._cust_map.get(cust_name)
                data = self.db.report_customer_ledger(cust_id, df, dt)
            elif rpt["key"] == "slow_moving":      data = self.db.report_slow_moving()
            elif rpt["key"] == "supplier_payables": data = self.db.report_supplier_payables()
            elif rpt["key"] == "customer_ageing":   data = self.db.report_customer_ageing()
            else:
                data = []
        except Exception as e:
            messagebox.showerror("Error", str(e), parent=self.winfo_toplevel())
            return

        self._report_data = data
        self._render_table(rpt, data)

    def _render_table(self, rpt: dict, data: list):
        # Rebuild columns
        self.tree.delete(*self.tree.get_children())
        cols = [c[0] for c in rpt["cols"]]
        self.tree.configure(columns=cols)
        for key, head, w in rpt["cols"]:
            self.tree.heading(key, text=head)
            self.tree.column(key, width=w, minwidth=50)

        # Row colouring for certain reports
        danger_keys  = {"out_of_stock", "Out of Stock"}
        warning_keys = {"Low Stock"}

        total = 0.0
        summary_col = rpt.get("summary_col")

        for i, row in enumerate(data):
            vals = [str(row.get(c[0], "")) for c in rpt["cols"]]

            # Pick tag
            tag = ""
            if rpt["key"] == "low_stock":
                tag = "danger" if row.get("status") == "Out of Stock" else "warning"
            elif rpt["key"] == "customer_ledger":
                tag = "danger" if row.get("txn_type") == "Credit" else "alt"
            else:
                tag = "alt" if i % 2 == 0 else ""

            self.tree.insert("", "end", iid=str(i), values=vals, tags=(tag,))

            if summary_col:
                try:
                    total += float(row.get(summary_col) or 0)
                except Exception:
                    pass

        # Summary line
        if summary_col and data:
            self._summary_lbl.configure(
                text=f"Total rows: {len(data)}    |    "
                     f"Grand Total ({rpt['cols'][[ c[0] for c in rpt['cols']].index(summary_col)][1]}): "
                     f"₹{total:,.2f}"
            )
        else:
            self._summary_lbl.configure(text=f"Total rows: {len(data)}")

    # ── Excel export ──────────────────────────────────────────
    def _export_excel(self):
        if not self._report_data or not self._report_def:
            messagebox.showinfo("No Data", "Generate a report first.",
                                parent=self.winfo_toplevel())
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror(
                "Missing Library",
                "openpyxl is required for Excel export.\n\n"
                "Run:  pip install openpyxl",
                parent=self.winfo_toplevel())
            return

        rpt  = self._report_def
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel file", "*.xlsx")],
            initialfile=f"{rpt['key']}.xlsx",
            parent=self.winfo_toplevel()
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = rpt["title"]

        hdr_font = Font(bold=True, color="FFFFFF", size=12)
        hdr_fill = PatternFill("solid", fgColor="1565C0")
        headers  = [c[1] for c in rpt["cols"]]
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(self._report_data, 2):
            for ci, (key, _, _) in enumerate(rpt["cols"], 1):
                ws.cell(row=ri, column=ci, value=row.get(key, ""))

        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(path)
        messagebox.showinfo("Exported", f"Saved to:\n{path}", parent=self.winfo_toplevel())
        try:
            os.startfile(path)
        except Exception:
            pass

    # ── CSV export ────────────────────────────────────────────
    def _export_csv(self):
        if not self._report_data or not self._report_def:
            messagebox.showinfo("No Data", "Generate a report first.",
                                parent=self.winfo_toplevel())
            return

        rpt  = self._report_def
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV file", "*.csv")],
            initialfile=f"{rpt['key']}.csv",
            parent=self.winfo_toplevel()
        )
        if not path:
            return

        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([c[1] for c in rpt["cols"]])
            for row in self._report_data:
                writer.writerow([row.get(c[0], "") for c in rpt["cols"]])

        messagebox.showinfo("Exported", f"Saved to:\n{path}",
                            parent=self.winfo_toplevel())
        try:
            import os
            os.startfile(path)
        except Exception:
            pass

    # ── PDF export ────────────────────────────────────────────
    def _export_pdf(self):
        if not self._report_data or not self._report_def:
            messagebox.showwarning("No Data",
                                   "Generate a report first.",
                                   parent=self.winfo_toplevel())
            return
        rpt  = self._report_def
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF file", "*.pdf")],
            initialfile=f"{rpt['key']}.pdf",
            parent=self.winfo_toplevel()
        )
        if not path:
            return
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.units import mm
            from reportlab.lib import colors
            from reportlab.platypus import (SimpleDocTemplate, Table,
                                            TableStyle, Paragraph, Spacer)
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.enums import TA_CENTER, TA_LEFT

            BLUE  = colors.HexColor(COLORS["btn_primary"])
            WHITE = colors.white
            LGRAY = colors.HexColor("#F5F7FF")

            # Use landscape for wide reports
            ncols = len(rpt["cols"])
            psize = landscape(A4) if ncols > 6 else A4

            doc = SimpleDocTemplate(path, pagesize=psize,
                                    leftMargin=12*mm, rightMargin=12*mm,
                                    topMargin=10*mm,  bottomMargin=10*mm)

            hdr_style = ParagraphStyle("h", fontSize=14, fontName="Helvetica-Bold",
                                        textColor=BLUE, alignment=TA_CENTER)
            sub_style = ParagraphStyle("s", fontSize=9,  fontName="Helvetica",
                                        textColor=colors.HexColor(COLORS["text_muted"]),
                                        alignment=TA_CENTER, spaceAfter=6)
            cell_s    = ParagraphStyle("c", fontSize=8, fontName="Helvetica",
                                        textColor=colors.HexColor("#1A1A2E"))
            th_s      = ParagraphStyle("t", fontSize=8, fontName="Helvetica-Bold",
                                        textColor=WHITE)

            story = [
                Paragraph(rpt["title"], hdr_style),
                Paragraph(f"Generated: {__import__('datetime').datetime.now().strftime('%d %b %Y  %I:%M %p')}",
                          sub_style),
                Spacer(1, 4*mm),
            ]

            # Header row
            tbl_data = [[Paragraph(c[1], th_s) for c in rpt["cols"]]]
            for row in self._report_data:
                tbl_data.append([
                    Paragraph(str(row.get(c[0], "")), cell_s)
                    for c in rpt["cols"]
                ])

            # Auto column widths
            page_w = (psize[0] - 24*mm)
            cw = page_w / ncols
            col_widths = [cw] * ncols

            tbl = Table(tbl_data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                ("BACKGROUND",    (0,0), (-1,0),  BLUE),
                ("ROWBACKGROUNDS",(0,1), (-1,-1),  [WHITE, LGRAY]),
                ("GRID",          (0,0), (-1,-1),  0.3, colors.HexColor(COLORS["border"])),
                ("TOPPADDING",    (0,0), (-1,-1),  3),
                ("BOTTOMPADDING", (0,0), (-1,-1),  3),
                ("LEFTPADDING",   (0,0), (-1,-1),  4),
                ("RIGHTPADDING",  (0,0), (-1,-1),  4),
                ("VALIGN",        (0,0), (-1,-1),  "MIDDLE"),
            ]))
            story.append(tbl)

            if rpt.get("summary_col") and self._report_data:
                try:
                    total = sum(float(r.get(rpt["summary_col"], 0) or 0)
                                for r in self._report_data)
                    story.append(Spacer(1, 4*mm))
                    story.append(Paragraph(
                        f"<b>Total rows: {len(self._report_data)}  |  "
                        f"Grand Total ({rpt['summary_col']}): {total:,.2f}</b>",
                        ParagraphStyle("sum", fontSize=9, fontName="Helvetica-Bold",
                                       textColor=BLUE, alignment=TA_LEFT)
                    ))
                except Exception:
                    pass

            doc.build(story)
            messagebox.showinfo("PDF Exported", f"Saved to:\n{path}",
                                parent=self.winfo_toplevel())
            # Open PDF — try multiple methods (RPT-2 fix)
            try:
                import os; os.startfile(path)
            except Exception:
                try:
                    import subprocess; subprocess.Popen(["start", "", path], shell=True)
                except Exception:
                    try:
                        import webbrowser; webbrowser.open(path)
                    except Exception:
                        pass

        except ImportError:
            messagebox.showerror("Missing Library",
                                 "reportlab is required.\nRun: pip install reportlab",
                                 parent=self.winfo_toplevel())
        except Exception as e:
            messagebox.showerror("PDF Error", str(e),
                                 parent=self.winfo_toplevel())

    # ── Drill-down ────────────────────────────────────────────
    def _on_tree_drilldown(self, event=None):
        if not self._active_key or not self._report_data:
            return
        sel = self.tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        row_data = self._report_data[idx]
        if self._active_key == "daily_sales":
            self._show_daily_sales_detail(str(row_data.get("date", "")))

    def _show_daily_sales_detail(self, date_str: str):
        if not date_str:
            return
        bills = self.db.get_bills(date_from=date_str, date_to=date_str, status="Active")

        dlg = ctk.CTkToplevel(self.winfo_toplevel())
        dlg.title(f"Bills — {date_str}")
        place_popup(dlg, 840, 520, self.winfo_toplevel())
        dlg.grab_set()
        dlg.attributes("-topmost", True)

        # Header
        hdr = ctk.CTkFrame(dlg, fg_color=COLORS["btn_primary"], corner_radius=0, height=58)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        ctk.CTkLabel(hdr, text=f"📅  Bills for {date_str}",
                     font=FONTS["subheading"], text_color="white"
                    ).pack(side="left", padx=20, pady=10)
        active = [b for b in bills if b.get("status") == "Active"]
        grand  = sum(float(b.get("grand_total", 0)) for b in active)
        ctk.CTkLabel(hdr,
                     text=f"{len(active)} bill(s)   |   Total ₹{grand:,.2f}",
                     font=FONTS["body_bold"], text_color="#BFDBFE"
                    ).pack(side="right", padx=20)

        # Table
        frame = ctk.CTkFrame(dlg, fg_color=COLORS["bg_card"], corner_radius=16)
        frame.pack(fill="both", expand=True, padx=12, pady=8)
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        cols   = ("bill_no", "time", "customer", "total", "payment", "status")
        heads  = ("Bill #",  "Time", "Customer", "Total ₹", "Payment Mode", "Status")
        widths = (110, 75, 220, 110, 140, 80)
        tree = ttk.Treeview(frame, columns=cols, show="headings",
                             style="Rpt.Treeview", selectmode="browse")
        for col, head, w in zip(cols, heads, widths):
            tree.heading(col, text=head)
            anch = "e" if col == "total" else "w"
            tree.column(col, width=w, anchor=anch, minwidth=50)

        vsb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        tree.grid(row=0, column=0, sticky="nsew", padx=(6, 0), pady=6)
        vsb.grid(row=0, column=1, sticky="ns", pady=6)

        _row_colors = COLORS["ROW_COLORS"]
        for i, b in enumerate(bills):
            raw_dt   = str(b.get("bill_date", ""))
            time_str = raw_dt[11:16] if len(raw_dt) >= 16 else ""
            is_void  = b.get("status") == "Void"
            tag      = "void" if is_void else f"row{i % len(_row_colors)}"
            tree.insert("", "end", iid=str(b["bill_id"]), values=(
                b.get("bill_number", ""),
                time_str,
                b.get("customer_name", "Walk-in Customer"),
                f"{b.get('grand_total', 0):,.2f}",
                b.get("payment_mode", ""),
                b.get("status", "Active"),
            ), tags=(tag,))

        tree.tag_configure("void", background=COLORS["badge_void"])
        for idx, color in enumerate(_row_colors):
            tree.tag_configure(f"row{idx}", background=color)

        ctk.CTkButton(dlg, text="Close", font=FONTS["button"],
                      fg_color=COLORS["btn_secondary"], height=44, corner_radius=12,
                      command=dlg.destroy
                     ).pack(padx=20, pady=(0, 12))
